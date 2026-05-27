"""
fetch_ga4.py — Coleta métricas diárias do GA4 e grava no SharePoint (OneDrive/Excel).

Modos de operação:
  HISTÓRICO — primeira execução: coleta de 01/01/2025 até ontem, dia a dia
  DIÁRIO    — execuções seguintes: coleta só o dia anterior

Secrets necessários no GitHub Actions:
  GA4_CREDENTIALS_JSON   → conteúdo do JSON da service account do GA4
  AZURE_TENANT_ID        → tenant ID do Azure AD
  AZURE_CLIENT_ID        → client ID do App Registration
  AZURE_CLIENT_SECRET    → client secret do App Registration
  SHAREPOINT_SITE_ID     → ID do site SharePoint
  SHAREPOINT_FILE_ID     → ID do arquivo Excel no SharePoint
"""

import os
import json
import datetime
import io
import sys
import time

import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
import openpyxl

# ──────────────────────────────────────────────
# Configurações
# ──────────────────────────────────────────────

GA4_PROPERTY_ID  = "326912205"
HISTORY_START    = datetime.date(2025, 1, 1)
SHEET_NAME       = "analytics"

# Colunas da planilha — ordem fixa, não alterar sem migrar o arquivo
COLUMNS = [
    "data",
    # Volume
    "sessoes",
    "usuarios",
    "novos_usuarios",
    "pageviews",
    # Conversão
    "compras",
    "taxa_conversao_pct",
    # Engajamento
    "taxa_rejeicao_pct",
    "duracao_media_seg",
    # Dispositivo
    "sessoes_mobile",
    "sessoes_desktop",
    "sessoes_tablet",
    # Canal
    "sessoes_organico",
    "sessoes_direto",
    "sessoes_pago",
    "sessoes_social",
    "sessoes_email",
    "sessoes_referral",
    "sessoes_outros_canais",
    # Funil
    "funil_search",
    "funil_select_item",
    "funil_add_to_cart",
    "funil_begin_checkout",
    "funil_purchase",
    # Rotas — top 5 origens e destinos do dia
    "top_origem_1", "top_origem_1_sessoes",
    "top_origem_2", "top_origem_2_sessoes",
    "top_origem_3", "top_origem_3_sessoes",
    "top_origem_4", "top_origem_4_sessoes",
    "top_origem_5", "top_origem_5_sessoes",
    "top_destino_1", "top_destino_1_sessoes",
    "top_destino_2", "top_destino_2_sessoes",
    "top_destino_3", "top_destino_3_sessoes",
    "top_destino_4", "top_destino_4_sessoes",
    "top_destino_5", "top_destino_5_sessoes",
]

# ──────────────────────────────────────────────
# Autenticação GA4
# ──────────────────────────────────────────────

def get_ga4_service():
    creds_info = json.loads(os.environ["GA4_CREDENTIALS_JSON"])
    creds = service_account.Credentials.from_service_account_info(
        creds_info,
        scopes=["https://www.googleapis.com/auth/analytics.readonly"],
    )
    return build("analyticsdata", "v1beta", credentials=creds)


# ──────────────────────────────────────────────
# Autenticação Microsoft Graph
# ──────────────────────────────────────────────

def get_graph_token():
    url = (
        f"https://login.microsoftonline.com/"
        f"{os.environ['AZURE_TENANT_ID']}/oauth2/v2.0/token"
    )
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     os.environ["AZURE_CLIENT_ID"],
        "client_secret": os.environ["AZURE_CLIENT_SECRET"],
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    return resp.json()["access_token"]


# ──────────────────────────────────────────────
# Download / upload do Excel no SharePoint
# ──────────────────────────────────────────────

GRAPH_BASE = "https://graph.microsoft.com/v1.0"


def download_excel(token):
    site_id = os.environ["SHAREPOINT_SITE_ID"]
    file_id = os.environ["SHAREPOINT_FILE_ID"]
    url = f"{GRAPH_BASE}/sites/{site_id}/drive/items/{file_id}/content"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    return io.BytesIO(resp.content)


def upload_excel(token, excel_bytes: bytes):
    site_id = os.environ["SHAREPOINT_SITE_ID"]
    file_id = os.environ["SHAREPOINT_FILE_ID"]
    url = f"{GRAPH_BASE}/sites/{site_id}/drive/items/{file_id}/content"
    resp = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream",
        },
        data=excel_bytes,
    )
    resp.raise_for_status()
    print(f"  Upload OK — status {resp.status_code}")


# ──────────────────────────────────────────────
# Coleta GA4
# ──────────────────────────────────────────────

def run_report(service, date_str, dimensions, metrics, limit=10):
    body = {
        "dateRanges": [{"startDate": date_str, "endDate": date_str}],
        "dimensions": [{"name": d} for d in dimensions],
        "metrics":    [{"name": m} for m in metrics],
        "limit":      limit,
    }
    resp = (
        service.properties()
        .runReport(property=f"properties/{GA4_PROPERTY_ID}", body=body)
        .execute()
    )
    return resp.get("rows", [])


def safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def collect_metrics(service, date_str):
    metrics = {}

    # ── Métricas gerais ──────────────────────────────────────────────────────
    rows = run_report(
        service, date_str, [],
        ["sessions", "totalUsers", "newUsers", "transactions",
         "sessionConversionRate", "bounceRate",
         "averageSessionDuration", "screenPageViews"],
    )
    if rows:
        v = rows[0]["metricValues"]
        metrics["sessoes"]            = int(safe_float(v[0]["value"]))
        metrics["usuarios"]           = int(safe_float(v[1]["value"]))
        metrics["novos_usuarios"]     = int(safe_float(v[2]["value"]))
        metrics["compras"]            = int(safe_float(v[3]["value"]))
        metrics["taxa_conversao_pct"] = round(safe_float(v[4]["value"]) * 100, 4)
        metrics["taxa_rejeicao_pct"]  = round(safe_float(v[5]["value"]) * 100, 4)
        metrics["duracao_media_seg"]  = round(safe_float(v[6]["value"]), 2)
        metrics["pageviews"]          = int(safe_float(v[7]["value"]))
    else:
        for k in ["sessoes", "usuarios", "novos_usuarios", "compras",
                  "taxa_conversao_pct", "taxa_rejeicao_pct",
                  "duracao_media_seg", "pageviews"]:
            metrics[k] = 0

    # ── Dispositivos ─────────────────────────────────────────────────────────
    device_map = {"mobile": "sessoes_mobile", "desktop": "sessoes_desktop", "tablet": "sessoes_tablet"}
    for k in device_map.values():
        metrics[k] = 0
    rows = run_report(service, date_str, ["deviceCategory"], ["sessions"])
    for row in rows:
        dev = row["dimensionValues"][0]["value"].lower()
        key = device_map.get(dev)
        if key:
            metrics[key] = int(safe_float(row["metricValues"][0]["value"]))

    # ── Canais ───────────────────────────────────────────────────────────────
    channel_map = {
        "organic search": "sessoes_organico",
        "direct":         "sessoes_direto",
        "paid search":    "sessoes_pago",
        "organic social": "sessoes_social",
        "email":          "sessoes_email",
        "referral":       "sessoes_referral",
    }
    for k in channel_map.values():
        metrics[k] = 0
    metrics["sessoes_outros_canais"] = 0

    rows = run_report(service, date_str, ["sessionDefaultChannelGroup"], ["sessions"], limit=20)
    for row in rows:
        channel = row["dimensionValues"][0]["value"].lower()
        key = channel_map.get(channel)
        val = int(safe_float(row["metricValues"][0]["value"]))
        if key:
            metrics[key] = val
        else:
            metrics["sessoes_outros_canais"] += val

    # ── Funil ────────────────────────────────────────────────────────────────
    funil_map = {
        "search":         "funil_search",
        "select_item":    "funil_select_item",
        "add_to_cart":    "funil_add_to_cart",
        "begin_checkout": "funil_begin_checkout",
        "purchase":       "funil_purchase",
    }
    for k in funil_map.values():
        metrics[k] = 0
    rows = run_report(service, date_str, ["eventName"], ["eventCount"], limit=50)
    for row in rows:
        event = row["dimensionValues"][0]["value"]
        key = funil_map.get(event)
        if key:
            metrics[key] = int(safe_float(row["metricValues"][0]["value"]))

    # ── Rotas — top 5 origens ────────────────────────────────────────────────
    for i in range(1, 6):
        metrics[f"top_origem_{i}"]         = ""
        metrics[f"top_origem_{i}_sessoes"] = 0

    rows = run_report(
        service, date_str,
        ["customEvent:originCity"], ["sessions"],
        limit=5,
    )
    for i, row in enumerate(rows[:5], 1):
        city = row["dimensionValues"][0]["value"]
        sessions = int(safe_float(row["metricValues"][0]["value"]))
        if city and city != "(not set)":
            metrics[f"top_origem_{i}"]         = city
            metrics[f"top_origem_{i}_sessoes"] = sessions

    # ── Rotas — top 5 destinos ───────────────────────────────────────────────
    for i in range(1, 6):
        metrics[f"top_destino_{i}"]         = ""
        metrics[f"top_destino_{i}_sessoes"] = 0

    rows = run_report(
        service, date_str,
        ["customEvent:destinationCity"], ["sessions"],
        limit=5,
    )
    for i, row in enumerate(rows[:5], 1):
        city = row["dimensionValues"][0]["value"]
        sessions = int(safe_float(row["metricValues"][0]["value"]))
        if city and city != "(not set)":
            metrics[f"top_destino_{i}"]         = city
            metrics[f"top_destino_{i}_sessoes"] = sessions

    return metrics


# ──────────────────────────────────────────────
# Manipulação do Excel
# ──────────────────────────────────────────────

def ensure_sheet(wb):
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(COLUMNS)
        # Remove a aba padrão "Sheet" se existir e estiver vazia
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1:
            del wb["Sheet"]
    else:
        ws = wb[SHEET_NAME]
        if ws.max_row == 0 or ws.cell(1, 1).value != "data":
            ws.insert_rows(1)
            for i, col in enumerate(COLUMNS, 1):
                ws.cell(1, i).value = col
    return ws


def get_existing_dates(ws):
    dates = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            dates.add(str(row[0]))
    return dates


def append_row(ws, date_str, metrics):
    row = [date_str] + [metrics.get(col, 0) for col in COLUMNS[1:]]
    ws.append(row)


def workbook_to_bytes(wb):
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    today     = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)

    # ── Autenticar e baixar planilha ─────────────────────────────────────────
    print("Autenticando no Microsoft Graph...")
    token = get_graph_token()

    print("Baixando planilha do SharePoint...")
    excel_bytes = download_excel(token)
    wb = openpyxl.load_workbook(excel_bytes)
    ws = ensure_sheet(wb)

    existing_dates = get_existing_dates(ws)
    print(f"Datas já na planilha: {len(existing_dates)}")

    # ── Determinar modo de operação ──────────────────────────────────────────
    # Modo histórico: planilha vazia ou com menos de 5 linhas de dados
    historical_mode = len(existing_dates) < 5

    if historical_mode:
        print(f"MODO HISTÓRICO — coletando de {HISTORY_START} até {yesterday}")
        dates_to_collect = []
        current = HISTORY_START
        while current <= yesterday:
            date_str = current.strftime("%Y-%m-%d")
            if date_str not in existing_dates:
                dates_to_collect.append(date_str)
            current += datetime.timedelta(days=1)
    else:
        print(f"MODO DIÁRIO — coletando {yesterday}")
        date_str = yesterday.strftime("%Y-%m-%d")
        if date_str in existing_dates:
            print(f"Data {date_str} já existe. Nada a fazer.")
            sys.exit(0)
        dates_to_collect = [date_str]

    if not dates_to_collect:
        print("Nenhuma data nova para coletar.")
        sys.exit(0)

    # ── Coletar GA4 ──────────────────────────────────────────────────────────
    print("Autenticando no GA4...")
    ga4_service = get_ga4_service()

    rows_added = 0
    total = len(dates_to_collect)

    for i, date_str in enumerate(dates_to_collect, 1):
        print(f"  [{i}/{total}] {date_str}...", end=" ")
        try:
            metrics = collect_metrics(ga4_service, date_str)
            append_row(ws, date_str, metrics)
            rows_added += 1
            print(
                f"sessoes={metrics['sessoes']} "
                f"compras={metrics['compras']} "
                f"CR={metrics['taxa_conversao_pct']}%"
            )
        except Exception as e:
            print(f"ERRO: {e}")

        # No modo histórico, fazer upload a cada 30 dias coletados
        # para evitar perda de dados em caso de timeout
        if historical_mode and rows_added > 0 and rows_added % 30 == 0:
            print(f"  Salvando checkpoint ({rows_added} linhas)...")
            token = get_graph_token()  # renovar token
            upload_excel(token, workbook_to_bytes(wb))

        # Pequena pausa para não sobrecarregar a API do GA4
        if historical_mode:
            time.sleep(0.3)

    # ── Upload final ─────────────────────────────────────────────────────────
    if rows_added > 0:
        print(f"\nTotal de linhas adicionadas: {rows_added}")
        token = get_graph_token()
        print("Enviando planilha atualizada ao SharePoint...")
        upload_excel(token, workbook_to_bytes(wb))
    else:
        print("Nenhuma linha nova adicionada.")

    print("Concluído.")


if __name__ == "__main__":
    main()
